# excel.py

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
# from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.borders import (Border, Side)
import string

# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -


def to_excel(brandname, folder_loc, dict_of_df):
    """
    Creates an Excel workbook with seperate sheets for each DataFrame object
    (named accordingly). Each sheet also gets styling and any extra blank
    columns (those w/o a header value) removed.

    Parameters
    ----------
    brandname : str
        The brand name used for file naming.
    folder_loc : str
        The location / path of the directory to store the file in.
    dict_of_df : dict
        A dictionary of DataFrame objects to convert to Excel worksheets.
    """

    def build_sheets(dodf=dict_of_df):
        """
        Creates an Excel workbook with seperate sheets for each DataFrame.
        Adds the DataFrame data into each appropirate sheet.

        Parameters
        ----------
        dodf : dict, optional
            A dictionary of DataFrame objects used to build Excel worksheets,
            by default 'dict_of_df' passed in by 'to_excel'.

        Returns
        -------
        wkbk : openpyxl.Workbook()
            An Excel workbook object with necessary sheets/tabs created and
            the DataFrame data inserted into the appropriate sheets.
        """
        wkbk = openpyxl.Workbook()              # Create Workbook object
        default_sheets = wkbk.sheetnames        # List of default sheet names

        # [STEP 1] Create the necessary Sheets and name them.
        for i, key in enumerate(dodf.keys()):

            if (i < len(default_sheets)):
                sheet_name = default_sheets[i]
                active_sheet = wkbk[sheet_name]
            else:
                sheet_name = None

            # CASE: Change name of an existing sheet.
            if (sheet_name in default_sheets):
                active_sheet.title = key
            # CASE: No more default sheets. Add new sheet with a specific name.
            else:
                wkbk.create_sheet(index=i, title=key)

        # [STEP 2] Fill in the sheet with the formatted DataFrame.
        for key, df in dodf.items():
            # Select Sheet
            ws = wkbk[key]
            # Append each row of the DataFrame as a row in the Excel worksheet.
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)

        return wkbk
    # [END] build_sheets()

    def style_sheets(workbook, dodf=dict_of_df, rh=15, cw=18):
        """
        Applies styling to an Excel workbook. Changes the font, column header
        fill colors, and trims the extra columns containing no data. Row
        height and column width can also be specified (optional).

        Parameters
        ----------
        workbook : openpyxl.Workbook()
            An Excel workbook object that has already been processed with the
            'build_sheets' function.
        dict_of_df : dict, optional
            A dictionary of DataFrame objects, by default 'dict_of_df' passed
            into the function 'to_excel'.
        rh : int, optional
            Cell row height, by default 15.
        cw : int, optional
            Cell column width, by default 18.

        Returns
        -------
        worksheet: openpyxl.Workbook()
            An Excel sheet object that has been styled to be fancy.
        """
        COLUMN_WIDTH, ROW_HEIGHT = (cw, rh)
        FONT_STYLE = Font(name='Arial', size=10, bold=True)  # Arial 10pt Bold

        # Defining Color Constants:
        # See 'colors.png' for color name references.
        RED = 'FFFF6347'            # Tomato
        PURPLE = 'FF9370DB'         # MediumPurple
        DARKPURPLE = 'FF9932CC'     # DrakOrchid
        TEAL = 'FF48D1CC'           # MediumTurquoise
        GREEN = 'FF9ACD32'          # YellowGreen
        ORANGE = 'FFFFA500'         # Orange
        YELLOW = 'FFFFD700'         # Gold
        WHITE = 'FFFFFFFF'          # White
        # LIGHTBLUE = 'FF00FFFF'     # Aqua
        # LIME = 'FF7FFF00'          # Chartreuse

        # Specify the color fill of th header cells in each named sheet.
        color_dict = {"LINE SHEET": None,
                      "EDIT": None,
                      "MASTER IMPORT": {"A1": PURPLE,
                                        "B1": PURPLE,
                                        "C1": TEAL,
                                        "D1": GREEN,
                                        "E1": PURPLE,
                                        "F1": PURPLE,
                                        "G1": RED,
                                        "H1": RED,
                                        "I1": WHITE,
                                        "J1": WHITE,
                                        "K1": WHITE,
                                        "L1": WHITE,
                                        "M1": WHITE,
                                        "N1": WHITE,
                                        "O1": GREEN,
                                        "P1": YELLOW,
                                        "Q1": ORANGE
                                        },
                      "SKU KEY": {"A1": DARKPURPLE,
                                  "B1": PURPLE},
                      "IMAGE KEY": {"A1": PURPLE,
                                    "B1": PURPLE,
                                    "C1": PURPLE,
                                    "D1": PURPLE,
                                    "E1": ORANGE},
                      "IMAGE IMPORT": {"A1": PURPLE,
                                       "B1": PURPLE,
                                       "C1": PURPLE,
                                       "D1": PURPLE,
                                       "E1": ORANGE},
                      # "DEFAULTS": {},
                      # "VARIANTS": {}
                      }

        # Run through each sheet and style the font and colorfill.
        # key = the sheet names  |||  value = dict (or None).
        for key, value in color_dict.items():

            # CASE: No Color Styling Defined --> Skip & Go To Next Sheet.
            if (value is None):
                continue
            # CASE: Colorfill was specified --> Apply Styling Rules.
            else:
                # Select worksheet
                ws = workbook[key]
                # Select the first row
                first_row = ws['1:1']
                # An iterator for column references.
                col_iter = 1

                # [STEP 1] Apply Font! Delete Empty Columns!
                # Iterate through cells in the first row.
                for cell in first_row:
                    # CASE: No header value --> Delete column.
                    if (cell.value == ''):
                        ws.delete_cols()
                    # CASE: Has header value --> Apply font.
                    else:
                        cell.font = FONT_STYLE
                    col_iter += 1       # Increase iterator

                # [STEP 2] Apply Color Fill & Border to Headers!
                for cell_range, color_const in value.items():
                    # Store the cell's location.
                    cell = ws[cell_range]
                    # Apply the fill style to the cell.
                    cell.fill = PatternFill(start_color=color_const,
                                            end_color=color_const,
                                            fill_type="solid")
                    # Apply border
                    side = Side(style='thin')
                    cell.border = Border(left=side, right=side,
                                         top=side, bottom=side)

                # [STEP 3] Apply Column Width / Row Height!
                first_col = ws['A']
                num_of_cols = len(first_row)
                num_of_rows = len(first_col)
                # Columns:
                for col in string.ascii_uppercase[:num_of_cols]:
                    ws.column_dimensions[col].width = COLUMN_WIDTH
                # Rows:
                for row in range(1, num_of_rows+1):
                    ws.row_dimensions[row].height = ROW_HEIGHT

        return workbook
    # [END] style_sheets()

    def save_doc(wkbk, brandname=brandname, f_loc=folder_loc, ext='xlsx'):
        # Create Workbook Name & Save Location
        wkbk_nm = f"products - {brandname}"
        # Create Save Location
        save_loc = f"{f_loc}\\{wkbk_nm}.{ext}"
        # Save the workbook in the folder.
        wkbk.save(save_loc)
        # Return the location where we saved the workbook.
        return save_loc
    # [END] save_doc()

    # === RUNNER CODE === #
    built = build_sheets()
    styled = style_sheets(built)
    save_location = save_doc(styled)

    return save_location
